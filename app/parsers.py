"""Workbook parsing helpers shared by the discipline agents.

Two sheet shapes are supported:
  - key/value sheets: two columns, parameter name in col A, value in col B
  - table sheets: header row followed by data rows
"""

import io

from openpyxl import load_workbook


class ParseError(Exception):
    pass


def load_xlsx(data: bytes):
    try:
        return load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:
        raise ParseError(f"Could not read the file as .xlsx: {exc}") from exc


def read_kv_sheet(ws) -> dict:
    """Read a two-column parameter sheet into {name: value}."""
    out = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        key, value = row[0], row[1] if len(row) > 1 else None
        if key is None:
            continue
        out[str(key).strip()] = value
    return out


def read_table_sheet(ws) -> list[dict]:
    """Read a header + rows sheet into a list of row dicts.

    Adds "_row" (1-based worksheet row number) to each dict so findings can
    point at the exact row in the source document.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for idx, row in enumerate(rows[1:], start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record = {header[i]: row[i] for i in range(min(len(header), len(row))) if header[i]}
        record["_row"] = idx
        out.append(record)
    return out


def as_float(value) -> float | None:
    if value is None:
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None
