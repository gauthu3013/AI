"""Generate the three demo deliverables with deliberately seeded errors.

Seeded issues (what the agents should catch):
  Electrical
    - Earthing conductor 300 mm2 vs required 312.5 mm2 (undersized)
    - Touch potential 850 V vs tolerable 780 V limit
  Mechanical
    - CH-02 design pressure 6.0 bar below operating 6.5 bar
    - CRAH-03 missing rated cooling capacity
  Process
    - Duplicate tag RACK-B02
    - RACK-B04 missing unit load
  Cross-discipline (found by the digital twin)
    - IT load 2400 kW vs UPS capacity 2000 kW  -> critical power undersized
    - N+1 cooling violated (2 x 1200 kW chillers vs 2400 kW heat load)
    - CH-03 in the equipment list but has no mechanical data sheet

Run:  python samples/generate_samples.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

HERE = Path(__file__).resolve().parent
HEADER_FONT = Font(bold=True)


def _style_header(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT


def electrical():
    wb = Workbook()

    ws = wb.active
    ws.title = "Power System Summary"
    rows = [
        ("Parameter", "Value"),
        ("Utility Feed Capacity (MW)", 5.0),
        ("Transformer TX-01 Rating (MVA)", 3.15),
        ("Transformer TX-02 Rating (MVA)", 3.15),
        ("UPS UPS-A Rating (kW)", 1000),
        ("UPS UPS-B Rating (kW)", 1000),
        ("System Power Factor", 0.9),
    ]
    for row in rows:
        ws.append(row)
    _style_header(ws)
    ws.column_dimensions["A"].width = 38

    ws = wb.create_sheet("Earthing Calculation")
    rows = [
        ("Parameter", "Value"),
        ("Soil Resistivity (ohm-m)", 65),
        ("Fault Current (kA)", 25),
        ("Fault Duration (s)", 1.0),
        ("Material Constant k (A/mm2)", 80),
        ("Selected Conductor Cross-Section (mm2)", 300),  # seeded: required is 312.5
        ("Calculated Grid Resistance (ohm)", 0.48),
        ("Maximum Allowed Grid Resistance (ohm)", 1.0),
        ("Calculated Step Potential (V)", 610),
        ("Tolerable Step Potential (V)", 2210),
        ("Calculated Touch Potential (V)", 850),  # seeded: limit is 780
        ("Tolerable Touch Potential (V)", 780),
    ]
    for row in rows:
        ws.append(row)
    _style_header(ws)
    ws.column_dimensions["A"].width = 42

    wb.save(HERE / "electrical_earthing_and_power.xlsx")


def mechanical():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cooling Equipment"
    ws.append(("Tag", "Equipment", "Rated Cooling Capacity (kW)", "Design Temp (C)",
               "Operating Temp (C)", "Design Pressure (bar)", "Operating Pressure (bar)",
               "Redundancy Role"))
    rows = [
        ("CH-01", "Water-Cooled Chiller", 1200, 45, 38, 10.0, 6.5, "Duty"),
        ("CH-02", "Water-Cooled Chiller", 1200, 45, 38, 6.0, 6.5, "Duty"),  # seeded pressure error
        ("CRAH-01", "Computer Room Air Handler", 250, 40, 32, 4.0, 2.5, "Duty"),
        ("CRAH-02", "Computer Room Air Handler", 250, 40, 32, 4.0, 2.5, "Duty"),
        ("CRAH-03", "Computer Room Air Handler", None, 40, 32, 4.0, 2.5, "Standby"),  # seeded missing capacity
        ("CRAH-04", "Computer Room Air Handler", 250, 40, 32, 4.0, 2.5, "Standby"),
    ]
    for row in rows:
        ws.append(row)
    _style_header(ws)
    for col, width in zip("ABCDEFGH", (10, 28, 26, 16, 18, 20, 22, 18)):
        ws.column_dimensions[col].width = width
    wb.save(HERE / "mechanical_cooling_datasheets.xlsx")


def process():
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment List"
    ws.append(("Tag", "Description", "Category", "Area", "Quantity",
               "Unit Load (kW)", "Total Load (kW)", "Remarks"))
    rows = [
        ("RACK-A01", "GPU Rack Block A1", "IT", "Hall A", 10, 30, 300, "NVIDIA HGX"),
        ("RACK-A02", "GPU Rack Block A2", "IT", "Hall A", 10, 30, 300, ""),
        ("RACK-A03", "GPU Rack Block A3", "IT", "Hall A", 10, 30, 300, ""),
        ("RACK-A04", "GPU Rack Block A4", "IT", "Hall A", 10, 30, 300, ""),
        ("RACK-B01", "GPU Rack Block B1", "IT", "Hall B", 10, 30, 300, ""),
        ("RACK-B02", "GPU Rack Block B2", "IT", "Hall B", 10, 30, 300, ""),
        ("RACK-B02", "GPU Rack Block B3", "IT", "Hall B", 10, 30, 300, "duplicate tag (seeded)"),
        ("RACK-B04", "Storage Rack Block B4", "IT", "Hall B", 10, None, 300, "unit load missing (seeded)"),
        ("CH-01", "Water-Cooled Chiller", "Mechanical", "Plant Room", 1, None, None, ""),
        ("CH-02", "Water-Cooled Chiller", "Mechanical", "Plant Room", 1, None, None, ""),
        ("CH-03", "Water-Cooled Chiller (future)", "Mechanical", "Plant Room", 1, None, None, "no data sheet (seeded)"),
        ("CRAH-01", "Computer Room Air Handler", "Mechanical", "Hall A", 1, None, None, ""),
        ("CRAH-02", "Computer Room Air Handler", "Mechanical", "Hall A", 1, None, None, ""),
        ("CRAH-03", "Computer Room Air Handler", "Mechanical", "Hall B", 1, None, None, ""),
        ("CRAH-04", "Computer Room Air Handler", "Mechanical", "Hall B", 1, None, None, ""),
    ]
    for row in rows:
        ws.append(row)
    _style_header(ws)
    for col, width in zip("ABCDEFGH", (12, 30, 14, 12, 10, 16, 16, 28)):
        ws.column_dimensions[col].width = width
    wb.save(HERE / "process_it_equipment_list.xlsx")


if __name__ == "__main__":
    electrical()
    mechanical()
    process()
    print("Sample deliverables written to", HERE)
